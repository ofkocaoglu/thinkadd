import os
import json
from openpyxl import load_workbook

# Change working directory to script location
tool_dir = os.path.dirname(os.path.abspath(__file__))
os.chdir(tool_dir)

# File names
excel_filename = 'LPBF_Material_Universe.xlsx'
json_filename = 'materialData.json'

# Construct full paths
excel_path = os.path.join(tool_dir, excel_filename)
json_path = os.path.join(tool_dir, json_filename)

# Verify Excel file presence
if not os.path.exists(excel_path):
    raise FileNotFoundError(f"Excel file not found at {excel_path}")

# Load workbook and active sheet
wb = load_workbook(excel_path, data_only=True)
ws = wb.active

# ARGB code for commercial cells
commercial_color = 'FF00B0F0'

material_data = {}

# Iterate over columns based on header row
for col_cells in ws.iter_cols(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
    header = col_cells[0].value
    if not header:
        continue
    entries = []
    # Process rows under header
    for cell in col_cells[1:]:
        value = cell.value
        if value is None:
            continue
        # Determine cell fill color
        fill = getattr(cell.fill, 'start_color', cell.fill.fgColor)
        rgb = getattr(fill, 'rgb', None)
        is_comm = (rgb and rgb.upper() == commercial_color)
        # Append entry
        if is_comm:
            entries.append({'name': str(value).strip(), 'commercial': True})
        else:
            entries.append(str(value).strip())
    material_data[str(header).strip()] = entries

# Write JSON file
with open(json_path, 'w', encoding='utf-8') as f:
    json.dump(material_data, f, ensure_ascii=False, indent=2)

print(f"Generated {json_filename} with {len(material_data)} categories.")
